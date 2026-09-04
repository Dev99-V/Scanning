"""QC Phase 2 helper: đối chiếu full reference_stock với file Excel mẫu.
Đếm ĐỘNG từ file (không hardcode) — đúng policy state.json:reference_row_count_policy.
Exit 0 + in QC2_COMPARE_OK khi: count khớp, khớp từng dòng, TRIM sạch, không dòng thừa.
Usage: python3 qc_phase2_compare.py <xlsx_path> <db_url>
"""
import sys

import openpyxl
import psycopg2

xlsx_path, db_url = sys.argv[1], sys.argv[2]

wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
ws = wb["Stock Balance With Batch"]
header = [c.value for c in ws[5]]
assert header == ["Stock Code", "Warehouse", "CREATEDATE", "BATCH", "BIN", "Qty"], f"header row 5 changed: {header}"

expected = {}
for r in ws.iter_rows(min_row=6, max_col=6, values_only=True):
    if all(v is None or v == "" for v in r):
        continue
    sc, wh, _cd, b, bn, q = r
    expected[str(b)] = (str(sc).strip(), str(sc), str(wh).strip(), str(bn).strip(), str(bn), float(q))
print(f"expected data rows in file: {len(expected)}")

con = psycopg2.connect(db_url)
cur = con.cursor()
cur.execute("select batch_id, stock_code, stock_code_raw, warehouse, bin, bin_raw, qty from reference_stock")
rows = cur.fetchall()
con.close()
print(f"rows in reference_stock: {len(rows)}")

fails = []
if len(rows) != len(expected):
    fails.append(f"count mismatch db={len(rows)} file={len(expected)}")
seen = set()
for batch_id, sc, sc_raw, wh, bn, bn_raw, qty in rows:
    seen.add(batch_id)
    e = expected.get(batch_id)
    if e is None:
        fails.append(f"db row not in file: {batch_id}")
        continue
    esc, esc_raw, ewh, ebn, ebn_raw, eq = e
    if not (sc == esc and sc_raw == esc_raw and wh == ewh and bn == ebn and bn_raw == ebn_raw and float(qty) == eq):
        fails.append(f"value mismatch: {batch_id}")
    if sc != sc.strip() or bn != bn.strip():
        fails.append(f"untrimmed value: {batch_id}")
for b in set(expected) - seen:
    fails.append(f"file row missing in db: {b}")

if fails:
    print(f"FAILURES ({len(fails)}):")
    for f in fails[:10]:
        print(" -", f)
    sys.exit(1)
print("QC2_COMPARE_OK")
